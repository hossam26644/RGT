''' docstring'''
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from string import ascii_lowercase

class ExcelWriter():
    """docstring for ClassName"""
    def __init__(self):
        #self.wb = xlwt.Workbook()
        self.wb = Workbook()
        std = self.wb.get_sheet_by_name('Sheet')
        self.wb.remove_sheet(std)

    def add_table_to_sheet(self, table, sheet_name, header_list):

        ws = self.wb.create_sheet(sheet_name)
        ws.title = sheet_name
        self.write_table_header(ws, header_list)

        row = 2
        for key in table.keys(): #put elements
            ws.cell(row=row, column=1, value=key)
            values = table[key]
            if isinstance(values, list): #check if the value is a single value or a list
                for idx, val in enumerate(values):
                    ws.cell(row=row, column=2+idx, value=val)

            else:   
                    ws.cell(row=row, column=2, value=table[key])

            row+=1
        self.cols_adjust_size(ws, header_list)

    def cols_adjust_size(self, ws, header_list):
        for idx, header in enumerate(header_list):
            size = len(header)+5
            col = chr(ord("a") + idx)
            ws.column_dimensions[col].width = size

   
    def save_file(self, file_name="DefaultName.xlsx"):
        self.wb.save(file_name)

    def write_table_header(self, ws, header_list):

        fill_style = PatternFill(fill_type='solid',
                start_color='FFFF00')
      
        for idx, col_name in enumerate(header_list): #put the table header
            cell = ws.cell(row=1, column=1+idx, value=col_name)
            cell.fill = fill_style
        

    @staticmethod	
    def write_to_excel(table, sheet_name="DefaultName", file_name="DefaultName.xls"):
        row = 1
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        for key in table.keys():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=table[key])
            row+=1
        wb.save(file_name)