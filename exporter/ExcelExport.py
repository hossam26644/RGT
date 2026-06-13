''' docstring'''
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from string import ascii_lowercase
import pandas as pd

class ExcelWriter():
    """docstring for ClassName"""
    def __init__(self):
        #self.wb = xlwt.Workbook()
        self.wb = Workbook()
        std = self.wb.get_sheet_by_name('Sheet')
        self.wb.remove_sheet(std)
        self.color_codes = {"red":"EC7063", "green":"ABEBC6" , "yellow":"F9E79F", "blue":"d6eaf8"}

    def add_table_to_sheet(self, table, sheet_name, header_list, color_table=None, colored_cell_index=None,
        align_right_index=[], color_odd_element_index=[]):

        ws = self.wb.create_sheet(sheet_name)
        ws.title = sheet_name
        self.write_table_header(ws, header_list)

        row = 2
        for key in table.keys():
            ws.cell(row=row, column=1, value=key)
            values = table[key]
            if isinstance(values, list):
                for idx, val in enumerate(values):
                    ws.cell(row=row, column=2+idx, value=val)
            else:
                ws.cell(row=row, column=2, value=table[key])

            if color_table is not None:
                self.apply_color_to_cells(ws, color_table[key], row)

            for idx in align_right_index:
                cell = ws.cell(row=row, column=idx)
                cell.alignment = cell.alignment.copy(horizontal='right')

            row += 1

        # Pad-align and color odd elements for specified columns
        for col_idx in color_odd_element_index:
            right = col_idx in align_right_index
            # Collect raw string values and their rows
            cells_data = []
            for r in range(2, row):
                cell = ws.cell(row=r, column=col_idx)
                cells_data.append((r, str(cell.value) if cell.value is not None else ""))

            if not cells_data:
                continue

            max_len = max(len(s) for _, s in cells_data)

            # Pad strings
            padded = []
            for r, s in cells_data:
                p = s.rjust(max_len) if right else s.ljust(max_len)
                padded.append((r, p))

            # Find most abundant character per position
            from collections import Counter
            most_common_per_pos = []
            for i in range(max_len):
                cnt = Counter(p[i] for _, p in padded)
                most_common_per_pos.append(cnt.most_common(1)[0][0])

            # Write rich text with red for non-majority characters
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            red_font = InlineFont(color="FF0000")

            for r, p in padded:
                segments = []
                cur_text = ""
                cur_is_red = None
                for i, ch in enumerate(p):
                    is_red = (ch != most_common_per_pos[i])
                    if cur_is_red is None:
                        cur_is_red = is_red
                        cur_text = ch
                    elif is_red == cur_is_red:
                        cur_text += ch
                    else:
                        if cur_is_red:
                            segments.append(TextBlock(red_font, cur_text))
                        else:
                            segments.append(cur_text)
                        cur_text = ch
                        cur_is_red = is_red
                if cur_text:
                    if cur_is_red:
                        segments.append(TextBlock(red_font, cur_text))
                    else:
                        segments.append(cur_text)

                cell = ws.cell(row=r, column=col_idx)
                cell.value = CellRichText(*segments)
                if right:
                    cell.alignment = cell.alignment.copy(horizontal='right')

        self.cols_adjust_size(ws, header_list)

    def add_pd_to_sheet(self, pd_table, sheet_name):
        ws = self.wb.create_sheet(sheet_name)
        ws.title = sheet_name
        
        header_list = list(pd_table.columns)
        self.write_table_header(ws, header_list)
        
        for row_idx, row in enumerate(pd_table.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        
        self.cols_adjust_size(ws, header_list)

    def apply_color_to_cells(self,ws,sample_color_table,row):
        for colored_cell_index in sample_color_table:
            color = sample_color_table[colored_cell_index]
            fill_style = PatternFill(fill_type='solid',
                        start_color=self.color_codes[color])                
            cell = ws.cell(row=row, column=colored_cell_index)
            cell.fill = fill_style

    
    def cols_adjust_size(self, ws, header_list):
        for idx, header in enumerate(header_list):
            size = len(header)+5
            col = chr(ord("a") + idx)
            ws.column_dimensions[col].width = size

   
    def save_file(self, file_name="DefaultName.xlsx"):
        self.wb.save(file_name)

    def write_table_header(self, ws, header_list):

        fill_style = PatternFill(fill_type='solid',
                start_color=self.color_codes["blue"])
      
        for idx, col_name in enumerate(header_list): #put the table header
            cell = ws.cell(row=1, column=1+idx, value=col_name)
            cell.fill = fill_style
        

    @staticmethod	
    def write_to_excel(table, sheet_name="DefaultName", file_name="DefaultName.xls"):
        row = 1
        wb = Workbook()
        ws = wb.create_sheet(sheet_name)
        for key in table.keys():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=table[key])
            row+=1
        wb.save(file_name)